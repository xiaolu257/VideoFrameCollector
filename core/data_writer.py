# Project Path: core/data_writer.py
import os
import tempfile
from datetime import datetime

from PyQt6.QtWidgets import (
    QMessageBox
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from config import CELL_SIZE


def safe_save_workbook(wb, path):
    try:
        wb.save(path)
    except PermissionError:
        QMessageBox.warning(None, "文件被占用", f"请先关闭文件再保存:\n{path}")
        return False
    except Exception as e:
        QMessageBox.critical(None, "保存失败", f"保存文件时出错:\n{str(e)}")
        return False
    return True


def write_to_excel(data, output_folder, base_filename="文件信息汇总.xlsx"):
    # 构造唯一文件名，格式如 文件信息_20250718_153045.xlsx
    name_part, ext = os.path.splitext(base_filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_filename = f"{name_part}_{timestamp}{ext}"
    output_path = os.path.join(output_folder, unique_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "文件信息"

    headers = ['文件名', '路径', '类型', '大小(MB)', '时长', '关键词', '预览图']
    ws.append(headers)
    img_col_idx = headers.index('预览图') + 1

    for idx, item in enumerate(data, start=2):
        ws.append([item['文件名'], item['路径'], item['类型'], item['大小(MB)'],
                   item['时长'], item['关键词']])

        if item['预览图']:
            try:
                temp_img_path = os.path.join(tempfile.gettempdir(), f"preview_{idx}.png")
                with open(temp_img_path, 'wb') as f:
                    f.write(item['预览图'])
                img = XLImage(temp_img_path)
                img.anchor = f"{get_column_letter(img_col_idx)}{idx}"
                ws.add_image(img)
            except Exception as e:
                print("插入图片失败:", e)

        ws.row_dimensions[idx].height = CELL_SIZE

    success = safe_save_workbook(wb, output_path)
    if success:
        return output_path
    else:
        return None
